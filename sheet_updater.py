"""Excel 数据表更新模块"""
import openpyxl
import json
import os
import re
import shutil
from datetime import datetime
from config import (
    BASELINE_XLSX, OUTPUT_DIR, OUTPUT_PREFIX,
    COLUMN_MAP, CACHE_DIR, CHANGE_LOG_FILE
)


def ensure_output_dir():
    """确保输出目录存在"""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(CACHE_DIR, exist_ok=True)


def _resolve_newest_sheet() -> str:
    """解析最新数据表：outputs/ 下最近保存的 unit_data_v*.xlsx，若无则回退基准表。

    按文件修改时间取最新（版本号可能是 1.12 / 1.11.1.1a 等，不能按字典序比较），
    保证多次检查的变更在 outputs/ 中累积，而不是每次从基准表重建。
    """
    ensure_output_dir()
    candidates = [
        os.path.join(OUTPUT_DIR, f)
        for f in os.listdir(OUTPUT_DIR)
        if f.startswith(OUTPUT_PREFIX) and f.endswith(".xlsx")
    ]
    if not candidates:
        return BASELINE_XLSX
    candidates.sort(key=os.path.getmtime)
    return candidates[-1]


def load_workbook(path: str = None, data_only: bool = False) -> tuple:
    """加载 Excel 工作簿，返回 (workbook, sheet, row_map, col_map)

    path 为 None 时自动解析 outputs/ 下最新版本（实现跨版本变更累积），
    否则使用显式指定的文件。
    data_only=True 时读取公式的缓存计算结果（供 convert_to_json 导出）。
    注意：编辑/保存（balance_monitor）必须 data_only=False，否则公式会丢。
    row_map: {单位名: 行号}
    col_map: {列名: 列号}
    """
    path = path or _resolve_newest_sheet()
    wb = openpyxl.load_workbook(path, data_only=data_only)
    ws = wb.active

    # 建立列名→列号映射（第一行是表头）
    col_map = {}  # {列名: 列号(1-based)}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header and header in COLUMN_MAP:
            col_map[header] = col

    # 建立单位名→行号映射
    row_map = {}  # {单位名: 行号(1-based)}
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if name:
            row_map[str(name)] = row

    return wb, ws, row_map, col_map


def apply_change(ws, row_map: dict, col_map: dict,
                 unit: str, field: str, new_value: str) -> bool:
    """对工作表应用单条变更。返回是否成功。"""
    if unit not in row_map:
        print(f"  [SKIP] 未知单位: {unit}")
        return False

    cn_field = None
    for cn, en in COLUMN_MAP.items():
        if en == field:
            cn_field = cn
            break

    if cn_field not in col_map:
        # 尝试直接匹配中文列名
        if field in col_map:
            cn_field = field
        else:
            print(f"  [SKIP] 未知列名: {field}")
            return False

    row = row_map[unit]
    col = col_map[cn_field]

    old_val = ws.cell(row=row, column=col).value

    try:
        # 解析新值：可能是绝对值、百分比变化、或加减
        new_val = _compute_new_value(old_val, new_value)
    except Exception as e:
        print(f"  [ERROR] 计算新值失败: {e} | {unit}.{field}: {old_val} + {new_value}")
        return False

    ws.cell(row=row, column=col).value = new_val
    print(f"  [OK] {unit}.{field}: {old_val} → {new_val}")
    return True


def _compute_new_value(old_val, change_str: str):
    """根据变更描述计算新值"""
    s = str(change_str).strip()

    # 百分比: "+30%" / "-15%" / "30%" / "+30.5%"
    pct_match = re.match(r'^([+\-]?\d+\.?\d*)\s*%$', s)
    if pct_match and old_val is not None:
        pct = float(pct_match.group(1)) / 100.0
        old_num = float(str(old_val).replace(',', ''))
        return old_num * (1 + pct)

    # 加减: "+200" / "-50"
    add_match = re.match(r'^([+\-])(\d+\.?\d*)$', s)
    if add_match and old_val is not None:
        sign = 1 if add_match.group(1) == '+' else -1
        delta = float(add_match.group(2))
        old_num = float(str(old_val).replace(',', ''))
        return old_num + sign * delta

    # 绝对值: "300" / "300.5"
    abs_match = re.match(r'^(\d+\.?\d*)$', s)
    if abs_match:
        val = float(abs_match.group(1))
        if val == int(val):
            return int(val)
        return val

    # 无法解析：直接返回字符串
    return s


def save_new_sheet(wb, version: str) -> str:
    """保存新版 xlsx，返回文件路径"""
    ensure_output_dir()
    filename = f"{OUTPUT_PREFIX}{version}.xlsx"
    path = os.path.join(OUTPUT_DIR, filename)
    wb.save(path)
    print(f"\n[SAVED] {path}")
    return path


def log_changes(version: str, post_title: str, changes: list[dict]):
    """记录变更日志"""
    ensure_output_dir()
    log = []
    if os.path.exists(CHANGE_LOG_FILE):
        with open(CHANGE_LOG_FILE, "r", encoding="utf-8") as f:
            log = json.load(f)

    log.append({
        "version": version,
        "title": post_title,
        "date": datetime.now().isoformat(),
        "changes": changes,
    })

    with open(CHANGE_LOG_FILE, "w", encoding="utf-8") as f:
        json.dump(log, f, ensure_ascii=False, indent=2)

    print(f"[LOG] 变更日志已更新 ({len(changes)} 条变动)")


def copy_baseline() -> str:
    """复制基准表到输出目录作为初始版本"""
    ensure_output_dir()
    dest = os.path.join(OUTPUT_DIR, f"{OUTPUT_PREFIX}baseline.xlsx")
    shutil.copy2(BASELINE_XLSX, dest)
    print(f"[INIT] 基准表已复制到 {dest}")
    return dest


if __name__ == "__main__":
    # 测试：加载并显示表结构
    wb, ws, row_map, col_map = load_workbook()
    print(f"单位: {list(row_map.keys())[:5]}...")
    print(f"列: {list(col_map.keys())[:5]}...")

    # 测试变更应用
    test_changes = [
        {"unit": "爬虫", "field": "单体血量", "new": "300"},
    ]
    for c in test_changes:
        apply_change(ws, row_map, col_map, c["unit"], c["field"], c["new"])
    save_new_sheet(wb, "test")
